import tkinter as tk
import customtkinter as ctk
from tkinter import ttk, messagebox
import threading
import time
import os
os.environ['PYGAME_HIDE_SUPPORT_PROMPT'] = "hide"
import pygame
import sys
import serial
import random
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import logging

from camera import open_camera_window
from data_manager import load_data, save_data

log_folder = 'logs'
os.makedirs(log_folder, exist_ok=True)
logging.basicConfig(
    filename=os.path.join(log_folder, 'project_slotcar.log'),
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

class ScrollableRadiobuttonFrame(ctk.CTkScrollableFrame):
    def __init__(self, master, item_list, command=None, **kwargs):
        super().__init__(master, **kwargs)

        self.command = command
        self.radiobutton_variable = ctk.StringVar()
        self.radiobutton_list = []
        for i, item in enumerate(item_list):
            self.add_item(item)

    def add_item(self, item):
        radiobutton = ctk.CTkRadioButton(self, text=item, value=item, variable=self.radiobutton_variable)
        if self.command is not None:
            radiobutton.configure(command=self.command)
        radiobutton.grid(row=len(self.radiobutton_list), column=0, pady=(0, 10))
        self.radiobutton_list.append(radiobutton)

    def remove_item(self, item):
        for radiobutton in self.radiobutton_list:
            if item == radiobutton.cget("text"):
                radiobutton.destroy()
                self.radiobutton_list.remove(radiobutton)
                return

    def get_checked_item(self):
        return self.radiobutton_variable.get()

class SlotCarManager(ctk.CTk):
    def __init__(self, root):
        super().__init__()
        self.root = root
        self.root.title("Project Slotcar")

        self.drivers = load_data('drivers.json', 1)
        self.results = load_data('results.json', 1)
        settings = load_data('settings.json', 0)
        self.serial_port = settings.get('serial_port', 'COM3')
        self.early_start_penalty = settings.get('early_start_penalty', 2)
        self.results_table2 = None
        self.results_window = None
        self.overlay_label = None

        self.debounce_interval = 500  # milliseconds
        self.debounce_timer = None

        self.disqualified = False

        self.create_widgets()
        pygame.mixer.init()

        self.save_settings()
        logging.info("SlotCarManager initialized")

    def save_settings(self):
        settings = {
            'serial_port': self.serial_port,
            'early_start_penalty': self.early_start_penalty
        }
        save_data('settings.json', settings)

    def create_widgets(self):
        frame = ctk.CTkFrame(self.root, corner_radius=10)
        frame.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

        self.driver_label = ctk.CTkLabel(frame, text="Driver Name:")
        self.driver_label.grid(row=0, column=0, pady=5)

        self.driver_entry = ctk.CTkEntry(frame)
        self.driver_entry.grid(row=0, column=1, pady=5)

        self.add_driver_button = ctk.CTkButton(frame, text="Add Driver", command=self.add_driver)
        self.add_driver_button.grid(row=0, column=2, pady=5, padx=5)

        self.remove_driver_button = ctk.CTkButton(frame, text="Remove Driver", command=self.remove_driver, fg_color='#dc3545', text_color='white')
        self.remove_driver_button.grid(row=0, column=3, pady=5, padx=5)
        
        self.scrollable_radiobutton_frame = ScrollableRadiobuttonFrame(master=self, width=200,
                                                                       item_list=self.drivers)
        self.scrollable_radiobutton_frame.grid(row=1, column=0, columnspan=4, pady=10, sticky="nsew")

        self.laps_label = ctk.CTkLabel(frame, text="Number of Laps:")
        self.laps_label.grid(row=2, column=0, pady=5)

        self.start_race_button = ctk.CTkButton(frame, text="Start Race", command=self.start_race, fg_color='#28a745', text_color='white')
        self.start_race_button.grid(row=2, column=2, pady=5, padx=5)

        self.start_race_button = ctk.CTkButton(frame, text="Disqualify", command=self.disqualify, fg_color='#dc3545', text_color='white')
        self.start_race_button.grid(row=2, column=3, pady=5, padx=5)

        self.laps_entry = ctk.CTkEntry(frame)
        self.laps_entry.grid(row=2, column=1, pady=5)

        self.scrollable_radiobutton_frame = ScrollableRadiobuttonFrame(master=frame, width=200, item_list=self.drivers)
        self.scrollable_radiobutton_frame.grid(row=1, column=0, columnspan=4, pady=10, sticky="nsew")

        self.port_label = ctk.CTkLabel(frame, text="Serial Port:")
        self.port_label.grid(row=3, column=0, pady=5)

        self.port_entry = ctk.CTkEntry(frame)
        self.port_entry.insert(0, self.serial_port)
        self.port_entry.grid(row=3, column=1, pady=5)

        self.set_port_button = ctk.CTkButton(frame, text="Set Port", command=self.set_serial_port)
        self.set_port_button.grid(row=3, column=2, pady=5, padx=5)

        self.penalty_label = ctk.CTkLabel(frame, text="Early Start Penalty (s):")
        self.penalty_label.grid(row=4, column=0, pady=5)

        self.penalty_entry = ctk.CTkEntry(frame)
        self.penalty_entry.insert(0, str(self.early_start_penalty))
        self.penalty_entry.grid(row=4, column=1, pady=5)

        self.set_penalty_button = ctk.CTkButton(frame, text="Set Penalty", command=self.set_early_start_penalty)
        self.set_penalty_button.grid(row=4, column=2, pady=5, padx=5)

        self.countdown_label = ctk.CTkLabel(frame, text="", font=("Helvetica", 16))
        self.countdown_label.grid(row=5, column=0, columnspan=4, pady=10)

        self.results_button = ctk.CTkButton(frame, text="Show Results", command=self.show_results)
        self.results_button.grid(row=6, column=0, sticky="nsew", pady=10, padx=(10,10))

        self.font_size_slider = ctk.CTkSlider(frame, from_=10, to=100, command=self.on_slider_change)
        self.font_size_slider.set(10)  # Set default value
        self.font_size_slider.grid(row=6, column=1, sticky="nsew", pady=10, padx=(10,10))

        self.open_camera_button = ctk.CTkButton(frame, text="Open Camera", command=lambda: threading.Thread(target=open_camera_window).start())
        self.open_camera_button.grid(row=6, column=2, sticky="nsew", pady=10, padx=(10,10))

        self.results_table = ttk.Treeview(frame, columns=("Rank", "Driver", "Last Lap", "Best Lap"), show='headings', style="Custom.Treeview")
        self.results_table.heading("Rank", text="Pos")
        self.results_table.heading("Driver", text="Driver")
        self.results_table.heading("Last Lap", text="Last Lap (s)")
        self.results_table.heading("Best Lap", text="Best Lap (s)")
        self.results_table.column("Rank", anchor=tk.CENTER, width=50)
        self.results_table.column("Driver", anchor=tk.CENTER, width=250)
        self.results_table.column("Last Lap", anchor=tk.CENTER, width=150)
        self.results_table.column("Best Lap", anchor=tk.CENTER, width=150)
        self.results_table.grid(row=7, column=0, columnspan=4, pady=10, sticky="nsew")
        self.update_results_table()
        logging.info("Widgets created")

    def disqualify(self):
        self.disqualified = True
        logging.info("Driver disqualified")

    def set_serial_port(self):
        try:
            self.serial_port = self.port_entry.get()
            self.save_settings()
            messagebox.showinfo("Serial Port Set", f"Serial port set to {self.serial_port}")
            logging.info(f"Serial port set to {self.serial_port}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to set serial port: {str(e)}")
            logging.error(f"Failed to set serial port: {str(e)}")

    def set_early_start_penalty(self):
        try:
            self.early_start_penalty = int(self.penalty_entry.get())
            self.save_settings()
            messagebox.showinfo("Penalty Set", f"Early start penalty set to {self.early_start_penalty} seconds")
            logging.info(f"Early start penalty set to {self.early_start_penalty} seconds")
        except ValueError:
            messagebox.showerror("Error", "Invalid penalty value. Please enter an integer.")
            logging.error("Invalid penalty value entered")

    def add_driver(self):
        try:
            driver_name = self.driver_entry.get()
            if driver_name:
                if driver_name not in self.drivers:
                    self.drivers.append(driver_name)
                    self.scrollable_radiobutton_frame.add_item(driver_name)
                    save_data('drivers.json', self.drivers)
                    self.driver_entry.delete(0, tk.END)
                else:
                    messagebox.showerror("Error", "Driver name already exists.")
                    logging.warning(f"Attempted to add duplicate driver: {driver_name}")
            else:
                messagebox.showerror("Error", "Driver name cannot be empty.")
                logging.warning("Attempted to add empty driver name")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to add driver: {str(e)}")
            logging.error(f"Failed to add driver: {str(e)}")

    def remove_driver(self):
        try:
            selected_driver = self.scrollable_radiobutton_frame.get_checked_item()
            if selected_driver:
                driver_name = selected_driver
                confirm = messagebox.askyesno("Confirm Deletion", f"Are you sure you want to remove {driver_name}?")
                if confirm:
                    self.drivers.remove(driver_name)
                    self.scrollable_radiobutton_frame.remove_item(driver_name)
                    self.results = [result for result in self.results if result['driver'] != driver_name]
                    save_data('drivers.json', self.drivers)
                    save_data('results.json', self.results)
                    messagebox.showinfo("Success", f"Driver {driver_name} removed.")
                    self.update_results_table()
            else:
                messagebox.showerror("Error", "No driver selected.")
                logging.warning("Attempted to remove driver with no selection")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to remove driver: {str(e)}")
            logging.error(f"Failed to remove driver: {str(e)}")

    def get_number_of_times(self):
        try:
            laps = int(self.laps_entry.get())
            if laps > 0:
                return laps
            else:
                messagebox.showerror("Error", "Number of laps must be greater than zero.")
                logging.error("Number of laps must be greater than zero.")
                return None
        except ValueError:
            messagebox.showerror("Error", "Invalid number of laps. Please enter an integer.")
            logging.error("Invalid number of laps. Please enter an integer.")
            return None
        
    def show_overlay(self, text):
        if self.overlay_label:
            self.overlay_label.destroy()
        if self.results_table2:
            self.overlay_label = ctk.CTkLabel(self.results_table2, text=text, font=("Helvetica", 128, "bold"), fg_color='#ffffff', text_color='#000000')
            self.overlay_label.place(relx=0.5, rely=0.5, anchor=tk.CENTER)

    def hide_overlay(self):
        if self.overlay_label:
            self.overlay_label.destroy()
            self.overlay_label = None

    def start_race(self):
        try:
            selected_driver = self.scrollable_radiobutton_frame.get_checked_item()
            if selected_driver:
                driver = selected_driver
                laps = self.get_number_of_laps()
                if driver and laps:
                    threading.Thread(target=self.countdown, args=(5, driver, laps)).start()  # 5 second countdown
            else:
                logging.warning("Attempted to start race with no driver selected")
                messagebox.showerror("Error", "No driver selected.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to start race: {str(e)}")
            logging.error(f"Failed to start race: {str(e)}")

    def play_sound(self, second):
        try:
            sound_folder_path = self.get_data_path('sounds')
            sound_file = f"{sound_folder_path}/{second}.wav"
            pygame.mixer.Sound(sound_file).play()
        except Exception as e:
            print(f"Failed to play sound: {str(e)}")
            logging.error(f"Failed to play sound: {str(e)}")

    def get_data_path(self,relative_path):
        if getattr(sys, 'frozen', False):
            # The application is running in a frozen state (e.g., bundled with PyInstaller)
            base_path = sys._MEIPASS
        else:
            # The application is running in a normal Python environment
            base_path = os.path.dirname(__file__)

        return os.path.join(base_path, relative_path)

    def countdown(self, seconds, driver, laps):
        try:
            with serial.Serial(self.serial_port, 9600, timeout=1) as ser:
                next = seconds
                counter = 0
                self.disqualified = False
                for second in range((seconds + 1) * 100, 0, -1):
                    time.sleep(0.01)
                    check_early = False
                    if ser.in_waiting > 0:
                        line = ser.readline().decode('utf-8').strip()
                        if line == '1':
                            check_early = True
                    if check_early:
                        ser.close()
                        self.show_overlay("Early Start!")
                        self.countdown_label.configure(text="Early Start!")
                        self.play_sound(f"false_start/{random.randint(1, 3)}")
                        self.run_race(driver, laps, True, True)
                        return
                    counter = counter + 1
                    if counter >= 100 and next > 0:
                        counter = 0
                        self.show_overlay(next)
                        self.countdown_label.configure(text=f"Stage starts in: {next}")
                        self.play_sound(f"countdown/{next}")
                        next = next - 1
                self.show_overlay("Go!")
                self.play_sound("countdown/GO")
                ser.close()
                self.run_race(driver, laps, False, False)
        except Exception as e:
            messagebox.showerror("Error", f"Serial communication error: {str(e)}")
            logging.error(f"Serial communication error: {str(e)}")
            return False

    def run_race(self, driver, laps, early_start, count_first):
        try:
            self.countdown_label.configure(text="")
            lap_times = []
            lap_count = 0
            count_first_temp = count_first
            logging.info("Started race...")
                
            while lap_count < laps:
                lap_start = time.time()
                lap_end = None
                with serial.Serial(self.serial_port, 9600, timeout=1) as ser:
                    while True:
                        if ser.in_waiting > 0:
                            line = ser.readline().decode('utf-8').strip()
                            if line == '1':
                                if count_first_temp:
                                    lap_end = time.time()
                                    break
                                else:
                                    count_first_temp = True
                        if self.disqualified:
                            self.disqualified = False
                            self.play_sound(f"disqualified/{random.randint(1, 3)}")
                            self.countdown_label.configure(text="Disqualified")
                            logging.info("Disqualified")
                            return
                lap_time = lap_end - lap_start
                if lap_count == 0 and early_start:
                    lap_time += self.early_start_penalty
                lap_times.append(lap_time)
                lap_count += 1
                self.hide_overlay()
                driver_found = False
                for result in self.results:
                    if result['driver'] == driver:
                        result['last_time'] = lap_time
                        result['best_time'] = min(result['best_time'], lap_time)
                        driver_found = True
                        break
                if not driver_found:
                    self.results.append({'driver': driver, 'last_time': lap_time, 'best_time': lap_time})
                save_data('results.json', self.results)
                self.update_results_table()
                logging.info("Finished race...")
                
            save_data('results.json', self.results)
            self.update_results_table()
            self.countdown_label.configure(text="Finished")
            self.play_sound(f"well_done/{random.randint(1, 3)}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to run race: {str(e)}")
            logging.error(f"Failed to run race: {str(e)}")

    def update_results_table(self):
        try:
            self.results_table.delete(*self.results_table.get_children())
            sorted_results = sorted([result for result in self.results if 'best_time' in result], key=lambda x: x['best_time'])
            font_size = self.font_size_slider.get()
            row_height = int(font_size * 1.5)
            
            style = ttk.Style()
            style.configure("ResultsWindow.Treeview", rowheight=row_height)

            for index, result in enumerate(sorted_results, start=1):
                self.results_table.insert("", tk.END, values=(index, result['driver'], f"{result['last_time']:.3f}", f"{result['best_time']:.3f}"), tags=('oddrow' if index % 2 == 0 else 'evenrow'))
                self.results_table.tag_configure('oddrow', background='white')
                self.results_table.tag_configure('evenrow', background='#f0f0f0')
            
            if self.results_window:
                try:
                    self.results_table2.delete(*self.results_table2.get_children())
                    sorted_results2 = sorted([result for result in self.results if 'best_time' in result], key=lambda x: x['best_time'])
                    for index, result in enumerate(sorted_results2, start=1):
                        self.results_table2.insert("", tk.END, values=(index, result['driver'], f"{result['last_time']:.3f}", f"{result['best_time']:.3f}"), tags=('oddrow' if index % 2 == 0 else 'evenrow'))
                        self.results_table2.tag_configure('oddrow', background='white')
                        self.results_table2.tag_configure('evenrow', background='#f0f0f0')
                except Exception as e:
                    logging.error(f"Failed to update results table: {str(e)}")
            
            self.dump_leaderboard_to_excel()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update results table: {str(e)}")
            logging.error(f"Failed to update results table: {str(e)}")

    
    def get_number_of_laps(self):
        try:
            laps = int(self.laps_entry.get())
            if laps > 0:
                return laps
            else:
                messagebox.showerror("Error", "Number of laps must be greater than zero.")
                logging.warning("Invalid laps value entered")
                return None
        except ValueError:
            messagebox.showerror("Error", "Invalid number of laps. Please enter an integer.")
            logging.warning("Invalid laps data type entered")
            return None

    def dump_leaderboard_to_excel(self):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Leaderboard"

            headers = ["Rank", "Driver", "Last Lap (s)", "Best Lap (s)"]
            ws.append(headers)

            for col in ws.iter_cols(min_col=1, max_col=4, min_row=1, max_row=1):
                for cell in col:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = Border(
                        left=Side(border_style="thin"),
                        right=Side(border_style="thin"),
                        top=Side(border_style="thin"),
                        bottom=Side(border_style="thin")
                    )
                    ws.column_dimensions[cell.column_letter].width = 20
                ws.column_dimensions["A"].width = 10
                ws.column_dimensions["B"].width = 40

            sorted_results = sorted([result for result in self.results if 'best_time' in result], key=lambda x: x['best_time'])

            for index, result in enumerate(sorted_results, start=1):
                row_data = [index, result['driver'], f"{result['last_time']:.3f}", f"{result['best_time']:.3f}"]
                row_index = ws.max_row + 1
                ws.append(row_data)

                for col_index, cell_value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_index, column=col_index)
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = Border(
                        left=Side(border_style="thin"),
                        right=Side(border_style="thin"),
                        top=Side(border_style="thin"),
                        bottom=Side(border_style="thin")
                    )

                    if index % 2 == 0:
                        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

            wb.save("leaderboard.xlsx")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to dump leaderboard to Excel: {str(e)}")
            logging.error(f"Failed to dump leaderboard to Excel: {str(e)}")

    def show_results(self):
        if self.results_window and self.results_window.winfo_exists():
            self.results_window.lift()
            return
        
        self.results_window = ctk.CTkToplevel(self.root)
        self.results_window.title("Race Results")
        self.results_table2 = ttk.Treeview(self.results_window, columns=("Rank", "Driver", "Last Lap", "Best Lap"), show='headings', style="ResultsWindow.Treeview")
        self.results_table2.heading("Rank", text="Pos")
        self.results_table2.heading("Driver", text="Driver")
        self.results_table2.heading("Last Lap", text="Last Lap (s)")
        self.results_table2.heading("Best Lap", text="Best Lap (s)")
        self.results_table2.column("Rank", anchor=tk.CENTER, width=50)
        self.results_table2.column("Driver", anchor=tk.CENTER, width=200)
        self.results_table2.column("Last Lap", anchor=tk.CENTER, width=300)
        self.results_table2.column("Best Lap", anchor=tk.CENTER, width=150)
        self.results_table2.pack(pady=10, padx=10, fill=tk.BOTH, expand=True)
        self.update_results_table()
        logging.info("Results window created")

    def update_font_size(self, size):
        if self.results_window:
            #width = self.results_window.winfo_width()
            #height = self.results_window.winfo_height()
            #font_size = max(10, int(min(width, height) * 0.03))
            custom_font = ("Arial", size, "bold")
            style = ttk.Style()
            style.configure("ResultsWindow.Treeview.Heading", font=custom_font)
            style.configure("ResultsWindow.Treeview", font=custom_font)
            self.update_results_table()
    
    def on_slider_change(self, value):
        if self.debounce_timer:
            self.root.after_cancel(self.debounce_timer)

        self.debounce_timer = self.root.after(self.debounce_interval, self.update_font_size, int(float(value)))

    def export_results_to_excel(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Race Results"
        headers = ["Rank", "Driver", "Last Lap (s)", "Best Lap (s)"]
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center')
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
        sorted_results = sorted(self.results, key=lambda x: x['best_time'])
        for rank, result in enumerate(sorted_results, start=1):
            driver = result['driver']
            lap_times = ", ".join(f"{t:.2f}" for t in result['lap_times'])
            best_time = result.get("last_time", "")
            row = [rank, driver, lap_times, best_time]
            for col_num, value in enumerate(row, 1):
                cell = sheet.cell(radio_framectkrow=rank + 1, column=col_num, value=value)
                cell.alignment = Alignment(horizontal='center')
        file_path = "race_results.xlsx"
        workbook.save(file_path)
        messagebox.showinfo("Export Successful", f"Results exported to {file_path}")
        logging.info(f"Results exported to {file_path}")

    def on_close(self):
        pygame.mixer.quit()
        self.root.destroy()
        sys.exit()

if __name__ == "__main__":
    ctk.set_appearance_mode("System")
    ctk.set_default_color_theme("blue")
    root = ctk.CTk()
    style = ttk.Style()
    style.configure("Custom.Treeview", font=("Helvetica", 14), rowheight=30)
    style.configure("Custom.Treeview.Heading", font=("Helvetica", 16, "bold"))
    style.configure("Custom.TreeviewLarge", font=("Helvetica", 20), rowheight=40)
    style.configure("Custom.TreeviewLarge.Heading", font=("Helvetica", 24, "bold"))
    app = SlotCarManager(root)
    root.protocol("WM_DELETE_WINDOW", app.on_close)
    root.mainloop()
